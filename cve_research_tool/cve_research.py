"""Standalone CVE reference investigation tool.

This script intentionally does not import the main CVE Hunter program.  It
borrows the same source idea: NVD, CVE.org, local KB, nuclei, Exploit-DB,
imfht, and optional web search.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT_DIR / "待补充cve.xlsx"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "output"

CVE_RE = re.compile(r"CVE-\d{4}-\d{4,}", re.IGNORECASE)

DEFAULT_SOURCES = ("nvd", "cveorg", "epss", "kev", "local", "nuclei")
DEEP_SOURCES = ("exploitdb", "imfht", "web")
ALL_SOURCES = DEFAULT_SOURCES + DEEP_SOURCES

NVD_API = "https://services.nvd.nist.gov/rest/json/cves/2.0"
CVE_ORG_API = "https://cveawg.mitre.org/api/cve/{cve_id}"
EPSS_API = "https://api.first.org/data/v1/epss"
CISA_KEV_FEED = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
NUCLEI_RAW = "https://raw.githubusercontent.com/projectdiscovery/nuclei-templates/main/http/cves"
IMFHT_BASE = "https://cve.imfht.com"


@dataclass(frozen=True)
class CVEInput:
    cve_id: str
    row_number: int
    status: str
    sheet: str


def main() -> int:
    args = parse_args()
    load_project_env(ROOT_DIR / ".env")

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    records_dir = output_dir / "records"
    output_dir.mkdir(parents=True, exist_ok=True)
    records_dir.mkdir(parents=True, exist_ok=True)

    sources = parse_sources(args.sources, deep=args.deep)
    cve_inputs = read_cve_inputs(
        input_path=input_path,
        sheet_name=args.sheet,
        cve_column=args.cve_column,
        status_column=args.status_column,
        include_generated=args.include_generated,
    )

    if args.start > 1:
        cve_inputs = cve_inputs[args.start - 1 :]
    if args.limit:
        cve_inputs = cve_inputs[: args.limit]

    if not cve_inputs:
        print("没有找到需要处理的 CVE。")
        return 1

    print(f"输入文件: {input_path}")
    print(f"处理数量: {len(cve_inputs)}")
    print(f"数据源: {', '.join(sources)}")
    print(f"输出目录: {output_dir}")

    timeout = args.timeout if args.timeout is not None else float(os.getenv("REQUEST_TIMEOUT", "30"))
    client = ResearchClient(
        timeout=timeout,
        proxy=get_proxy(),
        nvd_api_key=os.getenv("NVD_API_KEY", ""),
        tavily_api_key=os.getenv("TAVILY_API_KEY", ""),
        max_web_results=args.max_web_results,
    )

    epss_map: dict[str, dict[str, Any]] = {}
    if "epss" in sources:
        epss_map = client.query_epss_batch([item.cve_id for item in cve_inputs])
        print(f"EPSS 批量记录: {len(epss_map)}")

    kev_map: dict[str, dict[str, Any]] = {}
    if "kev" in sources:
        kev_map = client.query_kev_catalog()
        print(f"CISA KEV 目录记录: {len(kev_map)}")

    records: list[dict[str, Any]] = []
    loop_sources = [s for s in sources if s not in {"epss", "kev"}]
    for idx, item in enumerate(cve_inputs, 1):
        record_path = records_dir / f"{item.cve_id}.json"
        record = load_cached_record(record_path) if not args.refresh else None
        missing_sources = missing_loop_sources(record, loop_sources)

        if record is None:
            record = new_record(item)
        else:
            record.setdefault("sources_queried", [])
            record["input"] = input_to_dict(item)

        queried_network = False
        if missing_sources:
            for source in missing_sources:
                try:
                    record[source] = client.query_source(source, item.cve_id)
                except Exception as exc:  # keep batch moving
                    record[source] = {"found": False, "error": str(exc)}
                    record.setdefault("errors", []).append(f"{source}: {exc}")
                if source not in record["sources_queried"]:
                    record["sources_queried"].append(source)
                if source not in {"local"}:
                    queried_network = True

        if "epss" in sources:
            record["epss"] = epss_map.get(item.cve_id, {"found": False})
            if "epss" not in record["sources_queried"]:
                record["sources_queried"].append("epss")
        if "kev" in sources:
            record["kev"] = kev_map.get(item.cve_id, {"known_exploited": False})
            if "kev" not in record["sources_queried"]:
                record["sources_queried"].append("kev")

        record["summary"] = build_summary(record)
        record["updated_at"] = now_iso()
        save_json(record_path, record)
        records.append(record)

        summary = record["summary"]
        cached_label = "缓存" if not missing_sources else "查询"
        print(
            f"[{idx}/{len(cve_inputs)}] {item.cve_id} {cached_label} "
            f"refs={summary['total_reference_count']} "
            f"score={summary['information_score']} "
            f"{summary['category']}"
        )

        if queried_network and args.delay > 0 and idx < len(cve_inputs):
            time.sleep(args.delay)

    records.sort(
        key=lambda rec: (
            rec["summary"]["information_score"],
            rec["summary"]["total_reference_count"],
            rec["summary"]["source_hit_count"],
        ),
        reverse=True,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"cve_research_{timestamp}.xlsx"
    csv_path = output_dir / f"cve_research_{timestamp}.csv"
    jsonl_path = output_dir / f"cve_research_{timestamp}.jsonl"
    latest_xlsx = output_dir / "cve_research_latest.xlsx"
    latest_csv = output_dir / "cve_research_latest.csv"

    write_xlsx(records, xlsx_path)
    write_xlsx(records, latest_xlsx)
    write_csv(records, csv_path)
    write_csv(records, latest_csv)
    write_jsonl(records, jsonl_path)

    print("")
    print(f"已生成: {xlsx_path}")
    print(f"已生成: {csv_path}")
    print(f"已生成: {jsonl_path}")
    print(f"最新结果快捷文件: {latest_xlsx}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="独立 CVE 参考信息调查、汇总、排序和分类工具",
    )
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="待读取的 xlsx 文件")
    parser.add_argument("--sheet", default="", help="工作表名称，默认使用第一个工作表")
    parser.add_argument("--cve-column", default="", help="CVE 编号列名、列字母或列号，默认自动识别")
    parser.add_argument("--status-column", default="", help="工作流状态列名、列字母或列号，默认自动识别")
    parser.add_argument(
        "--include-generated",
        action="store_true",
        help="包含已经标记为“是/已生成”的 CVE，默认只处理未生成项",
    )
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="结果输出目录")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 个，便于试跑")
    parser.add_argument("--start", type=int, default=1, help="从待处理列表的第 N 个开始")
    parser.add_argument(
        "--sources",
        default="",
        help=(
            "逗号分隔的数据源。可选: nvd,cveorg,epss,kev,local,nuclei,"
            "exploitdb,imfht,web,all。默认: nvd,cveorg,epss,kev,local,nuclei"
        ),
    )
    parser.add_argument("--deep", action="store_true", help="在默认源基础上增加 exploitdb,imfht,web")
    parser.add_argument("--refresh", action="store_true", help="忽略已有 records 缓存，重新查询")
    parser.add_argument("--delay", type=float, default=0.6, help="逐个 CVE 网络查询后的等待秒数")
    parser.add_argument("--timeout", type=float, default=None, help="HTTP 超时秒数，默认读取 .env 的 REQUEST_TIMEOUT")
    parser.add_argument("--max-web-results", type=int, default=5, help="web 搜索最多记录多少条结果")
    return parser.parse_args()


def parse_sources(value: str, *, deep: bool) -> list[str]:
    if value:
        raw = [part.strip().lower() for part in value.split(",") if part.strip()]
        if "all" in raw:
            return list(ALL_SOURCES)
        unknown = sorted(set(raw) - set(ALL_SOURCES))
        if unknown:
            raise SystemExit(f"未知数据源: {', '.join(unknown)}")
        return dedupe(raw)

    sources = list(DEFAULT_SOURCES)
    if deep:
        sources.extend(DEEP_SOURCES)
    return dedupe(sources)


def load_project_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    try:
        from dotenv import load_dotenv

        load_dotenv(env_path)
        return
    except Exception:
        pass

    for raw_line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def get_proxy() -> str | None:
    for key in ("HTTPS_PROXY", "HTTP_PROXY", "https_proxy", "http_proxy"):
        value = os.getenv(key)
        if value:
            return value
    return None


def read_cve_inputs(
    *,
    input_path: Path,
    sheet_name: str,
    cve_column: str,
    status_column: str,
    include_generated: bool,
) -> list[CVEInput]:
    if not input_path.exists():
        raise SystemExit(f"输入文件不存在: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [cell_to_text(value) for value in rows[0]]
    cve_idx = resolve_column(cve_column, headers, fallback_keywords=("cve",))
    if cve_idx is None:
        cve_idx = detect_cve_column(rows[:20])
    if cve_idx is None:
        raise SystemExit("无法识别 CVE 编号列，请使用 --cve-column 指定。")

    status_idx = resolve_column(
        status_column,
        headers,
        fallback_keywords=("工作流是否已生成", "是否已生成", "已生成", "status"),
    )

    seen: set[str] = set()
    items: list[CVEInput] = []
    for row_number, row in enumerate(rows[1:], 2):
        value = row[cve_idx] if cve_idx < len(row) else ""
        match = CVE_RE.search(cell_to_text(value))
        if not match:
            continue
        cve_id = match.group(0).upper()
        if cve_id in seen:
            continue
        seen.add(cve_id)

        status = ""
        if status_idx is not None and status_idx < len(row):
            status = cell_to_text(row[status_idx])
        if not include_generated and is_generated_status(status):
            continue

        items.append(
            CVEInput(
                cve_id=cve_id,
                row_number=row_number,
                status=status,
                sheet=worksheet.title,
            )
        )

    return items


def resolve_column(spec: str, headers: list[str], fallback_keywords: tuple[str, ...]) -> int | None:
    if spec:
        normalized_spec = normalize_header(spec)
        if re.fullmatch(r"[A-Za-z]+", spec):
            return column_letter_to_index(spec)
        if spec.isdigit():
            return max(0, int(spec) - 1)
        for idx, header in enumerate(headers):
            if normalize_header(header) == normalized_spec:
                return idx
        for idx, header in enumerate(headers):
            if normalized_spec and normalized_spec in normalize_header(header):
                return idx
        return None

    for keyword in fallback_keywords:
        normalized_keyword = normalize_header(keyword)
        for idx, header in enumerate(headers):
            if normalized_keyword and normalized_keyword in normalize_header(header):
                return idx
    return None


def detect_cve_column(rows: list[tuple[Any, ...]]) -> int | None:
    counts: dict[int, int] = {}
    for row in rows:
        for idx, value in enumerate(row):
            if CVE_RE.search(cell_to_text(value)):
                counts[idx] = counts.get(idx, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda item: item[1])[0]


def column_letter_to_index(value: str) -> int:
    result = 0
    for char in value.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_generated_status(status: str) -> bool:
    return normalize_header(status) in {"是", "yes", "y", "true", "1", "已生成", "完成", "done"}


class ResearchClient:
    def __init__(
        self,
        *,
        timeout: float,
        proxy: str | None,
        nvd_api_key: str,
        tavily_api_key: str,
        max_web_results: int,
    ) -> None:
        self.timeout = timeout
        self.proxy = proxy
        self.nvd_api_key = nvd_api_key
        self.tavily_api_key = tavily_api_key
        self.max_web_results = max_web_results
        self.client = httpx.Client(
            timeout=timeout,
            proxy=proxy,
            follow_redirects=True,
            headers={"User-Agent": "CVEReferenceResearch/1.0"},
        )

    def query_source(self, source: str, cve_id: str) -> dict[str, Any]:
        if source == "nvd":
            return self.query_nvd(cve_id)
        if source == "cveorg":
            return self.query_cveorg(cve_id)
        if source == "local":
            return query_local_kb(cve_id)
        if source == "nuclei":
            return self.query_nuclei(cve_id)
        if source == "exploitdb":
            return self.query_exploitdb(cve_id)
        if source == "imfht":
            return self.query_imfht(cve_id)
        if source == "web":
            return self.query_web(cve_id)
        raise ValueError(f"Unsupported source: {source}")

    def query_nvd(self, cve_id: str) -> dict[str, Any]:
        headers: dict[str, str] = {}
        if self.nvd_api_key:
            headers["apiKey"] = self.nvd_api_key
        data = self.get_json_with_retries(
            NVD_API,
            params={"cveId": cve_id},
            headers=headers,
            source="nvd",
        )
        vulnerabilities = data.get("vulnerabilities", [])
        if not vulnerabilities:
            return {"found": False, "reference_count": 0, "error": f"NVD 中未找到 {cve_id}"}

        cve_item = vulnerabilities[0].get("cve", {})
        descriptions = cve_item.get("descriptions", [])
        description = first_lang_value(descriptions, "en") or first_lang_value(descriptions, "")
        references = parse_nvd_references(cve_item.get("references", []))
        products = parse_nvd_products(cve_item.get("configurations", []))
        cvss = parse_nvd_cvss(cve_item.get("metrics", {}))
        cwes = parse_nvd_cwes(cve_item.get("weaknesses", []))
        return {
            "found": True,
            "source_identifier": cve_item.get("sourceIdentifier", ""),
            "published": cve_item.get("published", ""),
            "last_modified": cve_item.get("lastModified", ""),
            "vuln_status": cve_item.get("vulnStatus", ""),
            "description": description,
            "reference_count": len(references),
            "references": references,
            "affected_product_count": len(products),
            "affected_products_sample": products[:8],
            "cwe_ids": cwes,
            **cvss,
        }

    def query_cveorg(self, cve_id: str) -> dict[str, Any]:
        data = self.get_json_with_retries(CVE_ORG_API.format(cve_id=cve_id), source="cveorg")
        metadata = data.get("cveMetadata", {})
        containers = data.get("containers", {})
        cna = containers.get("cna", {})
        adp_items = containers.get("adp", []) or []

        descriptions = cna.get("descriptions", [])
        description = first_lang_value(descriptions, "en") or first_lang_value(descriptions, "")

        references = []
        references.extend(parse_cveorg_references(cna.get("references", []), "cna"))
        for idx, adp in enumerate(adp_items, 1):
            references.extend(parse_cveorg_references(adp.get("references", []), f"adp{idx}"))

        affected = parse_cveorg_affected(cna.get("affected", []))
        problem_types = parse_cveorg_problem_types(cna.get("problemTypes", []))
        return {
            "found": bool(metadata or cna),
            "state": metadata.get("state", ""),
            "assigner": metadata.get("assignerShortName", ""),
            "date_published": metadata.get("datePublished", ""),
            "date_updated": metadata.get("dateUpdated", ""),
            "description": description,
            "reference_count": len(references),
            "references": references,
            "affected_count": len(affected),
            "affected_sample": affected[:8],
            "problem_types": problem_types,
        }

    def query_epss_batch(self, cve_ids: list[str]) -> dict[str, dict[str, Any]]:
        result: dict[str, dict[str, Any]] = {}
        for chunk in chunks(cve_ids, 100):
            try:
                data = self.get_json_with_retries(
                    EPSS_API,
                    params={"cve": ",".join(chunk)},
                    source="epss",
                )
                for item in data.get("data", []):
                    cve = item.get("cve", "").upper()
                    result[cve] = {
                        "found": True,
                        "epss": to_float(item.get("epss")),
                        "percentile": to_float(item.get("percentile")),
                        "date": item.get("date", ""),
                    }
            except Exception as exc:
                for cve_id in chunk:
                    result[cve_id] = {"found": False, "error": str(exc)}
        return result

    def query_kev_catalog(self) -> dict[str, dict[str, Any]]:
        try:
            data = self.get_json_with_retries(CISA_KEV_FEED, source="kev")
        except Exception as exc:
            return {"__error__": {"known_exploited": False, "error": str(exc)}}

        result: dict[str, dict[str, Any]] = {}
        for item in data.get("vulnerabilities", []):
            cve_id = item.get("cveID", "").upper()
            if not cve_id:
                continue
            result[cve_id] = {
                "known_exploited": True,
                "vendor_project": item.get("vendorProject", ""),
                "product": item.get("product", ""),
                "vulnerability_name": item.get("vulnerabilityName", ""),
                "date_added": item.get("dateAdded", ""),
                "due_date": item.get("dueDate", ""),
                "required_action": item.get("requiredAction", ""),
                "notes": item.get("notes", ""),
                "catalog_version": data.get("catalogVersion", ""),
            }
        return result

    def query_nuclei(self, cve_id: str) -> dict[str, Any]:
        match = CVE_RE.fullmatch(cve_id)
        if not match:
            return {"found": False, "error": "CVE 编号格式错误"}
        year = cve_id[4:8]
        url = f"{NUCLEI_RAW}/{year}/{cve_id}.yaml"
        response = self.client.get(url)
        if response.status_code == 200 and response.text.strip():
            text = response.text
            return {
                "found": True,
                "url": url,
                "name": first_yaml_scalar(text, "name"),
                "severity": first_yaml_scalar(text, "severity"),
                "reference_count": len(extract_urls(text)),
                "references": [{"url": u, "source": "nuclei"} for u in extract_urls(text)],
            }
        if response.status_code == 404:
            return {"found": False, "url": url}
        return {"found": False, "url": url, "error": f"HTTP {response.status_code}"}

    def query_exploitdb(self, cve_id: str) -> dict[str, Any]:
        response = self.client.get(
            "https://www.exploit-db.com/search",
            params={"cve": cve_id.replace("CVE-", "")},
            headers={
                "User-Agent": "Mozilla/5.0",
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/json",
            },
        )
        if response.status_code != 200:
            return {"found": False, "error": f"HTTP {response.status_code}"}
        try:
            data = response.json()
        except Exception as exc:
            return {"found": False, "error": f"JSON 解析失败: {exc}"}
        records = data.get("data", []) or []
        results = []
        for item in records[:10]:
            exploit_id = str(item.get("id", "")).strip()
            if not exploit_id:
                continue
            description = item.get("description", "")
            if isinstance(description, list):
                description = " ".join(str(part) for part in description)
            results.append(
                {
                    "id": exploit_id,
                    "title": strip_html(str(description)).strip(),
                    "url": f"https://www.exploit-db.com/exploits/{exploit_id}",
                    "type": strip_html(str(item.get("type", ""))),
                    "platform": strip_html(str(item.get("platform", ""))),
                }
            )
        return {
            "found": bool(results),
            "result_count": len(records),
            "results": results,
        }

    def query_imfht(self, cve_id: str) -> dict[str, Any]:
        url = f"{IMFHT_BASE}/{cve_id}"
        response = self.client.get(url, headers={"User-Agent": "Mozilla/5.0"})
        if response.status_code == 404:
            return {"found": False, "url": url}
        if response.status_code != 200:
            return {"found": False, "url": url, "error": f"HTTP {response.status_code}"}
        text = strip_html(response.text)
        found = cve_id.upper() in text.upper() and len(text) > 200
        return {
            "found": found,
            "url": url,
            "content_length": len(text),
            "summary": compact_text(text, 350) if found else "",
            "reference_count": len(extract_urls(response.text)),
            "references": [{"url": u, "source": "imfht"} for u in extract_urls(response.text)],
        }

    def query_web(self, cve_id: str) -> dict[str, Any]:
        query = f"{cve_id} PoC exploit vulnerability reference"
        if self.tavily_api_key:
            results = self.search_tavily(query)
            if results:
                return {"found": True, "query": query, "result_count": len(results), "results": results}
        results = self.search_duckduckgo(query)
        return {"found": bool(results), "query": query, "result_count": len(results), "results": results}

    def search_tavily(self, query: str) -> list[dict[str, str]]:
        try:
            from tavily import TavilyClient

            client = TavilyClient(api_key=self.tavily_api_key)
            data = client.search(
                query=query,
                max_results=self.max_web_results,
                search_depth="advanced",
                include_raw_content=False,
            )
            return [
                {
                    "title": item.get("title", ""),
                    "url": item.get("url", ""),
                    "content": compact_text(item.get("content", ""), 300),
                }
                for item in data.get("results", [])
                if item.get("url")
            ]
        except Exception:
            return []

    def search_duckduckgo(self, query: str) -> list[dict[str, str]]:
        response = self.client.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query},
            headers={"User-Agent": "Mozilla/5.0"},
        )
        response.raise_for_status()
        results = []
        for match in re.finditer(
            r'<a rel="nofollow" class="result__a" href="([^"]+)"[^>]*>(.+?)</a>',
            response.text,
            re.DOTALL,
        ):
            url = html_unescape(match.group(1))
            title = strip_html(match.group(2))
            results.append({"title": title, "url": url, "content": ""})
            if len(results) >= self.max_web_results:
                break
        return results

    def get_json_with_retries(
        self,
        url: str,
        *,
        params: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
        source: str,
    ) -> dict[str, Any]:
        last_error: Exception | None = None
        for attempt in range(1, 4):
            try:
                response = self.client.get(url, params=params, headers=headers)
                if is_rate_limited(response):
                    delay = retry_delay(response, attempt)
                    print(f"{source} 触发限流，等待 {delay}s 后重试。")
                    time.sleep(delay)
                    continue
                response.raise_for_status()
                return response.json()
            except Exception as exc:
                last_error = exc
                if attempt < 3:
                    time.sleep(1.5 * attempt)
                    continue
        raise RuntimeError(f"{source} 查询失败: {last_error}")


def query_local_kb(cve_id: str) -> dict[str, Any]:
    year = cve_id[4:8]
    files = [
        ("local_kb_custom", ROOT_DIR / "poc_kb" / "custom" / year / f"{cve_id}.md"),
        ("local_kb_trickest", ROOT_DIR / "poc_kb" / "trickest-cve" / year / f"{cve_id}.md"),
    ]

    hits = []
    urls: list[str] = []
    github_urls: list[str] = []
    has_http_poc = False
    has_yaml = False
    content_summary = ""

    for source, path in files:
        if not path.exists():
            continue
        content = path.read_text(encoding="utf-8", errors="ignore")
        found_urls = extract_urls(content)
        urls.extend(found_urls)
        github_urls.extend([url for url in found_urls if "github.com" in urlparse(url).netloc.lower()])
        has_http_poc = has_http_poc or bool(
            re.search(r"(?mi)^(GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS)\s+\S+", content)
        )
        has_yaml = has_yaml or bool(re.search(r"(?mi)^\s*id\s*:", content))
        if not content_summary:
            content_summary = compact_text(strip_markdown(content), 350)
        hits.append({"source": source, "path": str(path), "size": path.stat().st_size})

    urls = dedupe_urls(urls)
    github_urls = dedupe_urls(github_urls)
    return {
        "found": bool(hits),
        "hits": hits,
        "has_http_poc": has_http_poc,
        "has_yaml": has_yaml,
        "reference_count": len(urls),
        "references": [{"url": url, "source": "local_kb"} for url in urls],
        "github_repo_count": len(github_urls),
        "github_repos": github_urls[:20],
        "summary": content_summary,
    }


def new_record(item: CVEInput) -> dict[str, Any]:
    return {
        "cve_id": item.cve_id,
        "input": input_to_dict(item),
        "created_at": now_iso(),
        "updated_at": "",
        "sources_queried": [],
        "errors": [],
    }


def input_to_dict(item: CVEInput) -> dict[str, Any]:
    return {
        "row_number": item.row_number,
        "status": item.status,
        "sheet": item.sheet,
    }


def load_cached_record(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def missing_loop_sources(record: dict[str, Any] | None, sources: list[str]) -> list[str]:
    if not record:
        return sources
    queried = set(record.get("sources_queried", []))
    return [source for source in sources if source not in queried or source not in record]


def build_summary(record: dict[str, Any]) -> dict[str, Any]:
    cve_id = record["cve_id"]
    nvd = record.get("nvd", {})
    cveorg = record.get("cveorg", {})
    epss = record.get("epss", {})
    kev = record.get("kev", {})
    local = record.get("local", {})
    nuclei = record.get("nuclei", {})
    exploitdb = record.get("exploitdb", {})
    imfht = record.get("imfht", {})
    web = record.get("web", {})

    references = collect_references(record)
    total_reference_count = len(references)
    source_hits = []
    for source in ("nvd", "cveorg", "local", "nuclei", "exploitdb", "imfht", "web"):
        data = record.get(source, {})
        if data.get("found"):
            source_hits.append(source)
    if epss.get("found"):
        source_hits.append("epss")
    if kev.get("known_exploited"):
        source_hits.append("kev")

    poc_hint_count = 0
    if local.get("has_http_poc") or local.get("has_yaml"):
        poc_hint_count += 1
    if local.get("github_repo_count", 0):
        poc_hint_count += int(local.get("github_repo_count", 0))
    if nuclei.get("found"):
        poc_hint_count += 1
    if exploitdb.get("result_count", 0):
        poc_hint_count += int(exploitdb.get("result_count", 0))
    if web.get("result_count", 0):
        poc_hint_count += min(3, int(web.get("result_count", 0)))

    description = first_non_empty(
        nvd.get("description"),
        cveorg.get("description"),
        local.get("summary"),
        imfht.get("summary"),
    )
    affected_count = max(
        int(nvd.get("affected_product_count", 0) or 0),
        int(cveorg.get("affected_count", 0) or 0),
    )
    cwe_ids = dedupe([*nvd.get("cwe_ids", []), *cveorg.get("problem_types", [])])
    cvss_score = to_float(nvd.get("cvss_score"))
    cvss_severity = nvd.get("cvss_severity", "")

    information_score = score_information(
        has_description=bool(description),
        total_reference_count=total_reference_count,
        source_hit_count=len(source_hits),
        poc_hint_count=poc_hint_count,
        affected_count=affected_count,
        cwe_count=len(cwe_ids),
        has_cvss=cvss_score > 0,
    )
    priority_score = score_priority(
        information_score=information_score,
        cvss_score=cvss_score,
        epss_percentile=to_float(epss.get("percentile")),
        kev_known=bool(kev.get("known_exploited")),
        poc_hint_count=poc_hint_count,
    )
    category = classify_information(information_score, total_reference_count, len(source_hits))
    action = suggest_action(
        category=category,
        poc_hint_count=poc_hint_count,
        total_reference_count=total_reference_count,
        kev_known=bool(kev.get("known_exploited")),
        nuclei_found=bool(nuclei.get("found")),
        local_found=bool(local.get("found")),
        exploitdb_count=int(exploitdb.get("result_count", 0) or 0),
    )

    return {
        "cve_id": cve_id,
        "category": category,
        "suggested_action": action,
        "information_score": information_score,
        "priority_score": priority_score,
        "total_reference_count": total_reference_count,
        "nvd_reference_count": int(nvd.get("reference_count", 0) or 0),
        "cveorg_reference_count": int(cveorg.get("reference_count", 0) or 0),
        "local_reference_count": int(local.get("reference_count", 0) or 0),
        "web_result_count": int(web.get("result_count", 0) or 0),
        "source_hit_count": len(source_hits),
        "source_hits": source_hits,
        "poc_hint_count": poc_hint_count,
        "cvss_score": cvss_score,
        "cvss_severity": cvss_severity,
        "epss": to_float(epss.get("epss")),
        "epss_percentile": to_float(epss.get("percentile")),
        "kev_known": bool(kev.get("known_exploited")),
        "published": first_non_empty(nvd.get("published"), cveorg.get("date_published")),
        "last_modified": first_non_empty(nvd.get("last_modified"), cveorg.get("date_updated")),
        "vuln_status": first_non_empty(nvd.get("vuln_status"), cveorg.get("state")),
        "cwe_ids": cwe_ids,
        "affected_count": affected_count,
        "affected_sample": affected_sample(nvd, cveorg),
        "description_summary": compact_text(description, 360),
        "top_references": [ref["url"] for ref in references[:8]],
        "errors": collect_errors(record),
    }


def score_information(
    *,
    has_description: bool,
    total_reference_count: int,
    source_hit_count: int,
    poc_hint_count: int,
    affected_count: int,
    cwe_count: int,
    has_cvss: bool,
) -> int:
    score = 0
    if has_description:
        score += 8
    if has_cvss:
        score += 2
    score += min(total_reference_count, 30) * 3
    score += source_hit_count * 5
    score += min(poc_hint_count, 8) * 8
    score += min(affected_count, 12)
    score += min(cwe_count, 5)
    return int(score)


def score_priority(
    *,
    information_score: int,
    cvss_score: float,
    epss_percentile: float,
    kev_known: bool,
    poc_hint_count: int,
) -> float:
    score = float(information_score)
    if cvss_score:
        score += cvss_score * 2
    if epss_percentile:
        score += epss_percentile * 10
    if kev_known:
        score += 30
    score += min(poc_hint_count, 5) * 4
    return round(score, 2)


def classify_information(score: int, references: int, source_hits: int) -> str:
    if score >= 75 or references >= 18 or source_hits >= 6:
        return "A_信息丰富"
    if score >= 42 or references >= 8 or source_hits >= 4:
        return "B_信息中等"
    if score >= 18 or references >= 3 or source_hits >= 2:
        return "C_信息较少"
    if score > 0 or references > 0 or source_hits > 0:
        return "D_信息稀少"
    return "E_查询失败或空白"


def suggest_action(
    *,
    category: str,
    poc_hint_count: int,
    total_reference_count: int,
    kev_known: bool,
    nuclei_found: bool,
    local_found: bool,
    exploitdb_count: int,
) -> str:
    if kev_known:
        return "优先人工调查：CISA KEV 已知利用"
    if local_found:
        return "优先补剧本：本地知识库已有线索"
    if nuclei_found:
        return "优先补剧本：nuclei 模板可参考"
    if exploitdb_count:
        return "优先补剧本：Exploit-DB 有结果"
    if poc_hint_count >= 2:
        return "优先调查：PoC/GitHub 线索较多"
    if total_reference_count >= 8:
        return "先读参考链接：资料较充分"
    if category.startswith("E_"):
        return "低信息：需要人工扩展搜索或暂缓"
    if total_reference_count <= 2:
        return "信息少：先做厂商/关键词扩展搜索"
    return "常规调查"


def collect_references(record: dict[str, Any]) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    for source in ("nvd", "cveorg", "local", "nuclei", "imfht"):
        data = record.get(source, {})
        for ref in data.get("references", []) or []:
            url = ref.get("url", "")
            if url:
                refs.append(
                    {
                        "cve_id": record["cve_id"],
                        "source": source,
                        "url": url,
                        "label": ref.get("name", "") or ref.get("source", "") or ref.get("tags", ""),
                    }
                )
    for result in record.get("exploitdb", {}).get("results", []) or []:
        if result.get("url"):
            refs.append(
                {
                    "cve_id": record["cve_id"],
                    "source": "exploitdb",
                    "url": result["url"],
                    "label": result.get("title", ""),
                }
            )
    for result in record.get("web", {}).get("results", []) or []:
        if result.get("url"):
            refs.append(
                {
                    "cve_id": record["cve_id"],
                    "source": "web",
                    "url": result["url"],
                    "label": result.get("title", ""),
                }
            )
    return dedupe_reference_dicts(refs)


def collect_errors(record: dict[str, Any]) -> list[str]:
    errors = list(record.get("errors", []) or [])
    for source in ALL_SOURCES:
        data = record.get(source, {})
        if isinstance(data, dict) and data.get("error"):
            errors.append(f"{source}: {data['error']}")
    return dedupe(errors)


def affected_sample(nvd: dict[str, Any], cveorg: dict[str, Any]) -> list[str]:
    values = []
    values.extend(nvd.get("affected_products_sample", []) or [])
    values.extend(cveorg.get("affected_sample", []) or [])
    return dedupe([compact_product(value) for value in values if value])[:8]


def parse_nvd_references(items: list[dict[str, Any]]) -> list[dict[str, str]]:
    refs = []
    for item in items:
        url = item.get("url", "")
        if not url:
            continue
        refs.append(
            {
                "url": url,
                "source": item.get("source", ""),
                "tags": ",".join(item.get("tags", []) or []),
            }
        )
    return dedupe_reference_dicts(refs)


def parse_nvd_products(configurations: list[dict[str, Any]]) -> list[str]:
    products: list[str] = []

    def visit_node(node: dict[str, Any]) -> None:
        for match in node.get("cpeMatch", []) or []:
            criteria = match.get("criteria", "")
            if criteria:
                products.append(criteria)
        for child in node.get("nodes", []) or []:
            visit_node(child)

    for config in configurations or []:
        for node in config.get("nodes", []) or []:
            visit_node(node)
    return dedupe(products)


def parse_nvd_cvss(metrics: dict[str, Any]) -> dict[str, Any]:
    for key in ("cvssMetricV40", "cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        metric_list = metrics.get(key, []) or []
        if not metric_list:
            continue
        metric = metric_list[0]
        cvss_data = metric.get("cvssData", {})
        return {
            "cvss_version": cvss_data.get("version", key.replace("cvssMetric", "")),
            "cvss_score": to_float(cvss_data.get("baseScore", metric.get("baseScore"))),
            "cvss_severity": cvss_data.get("baseSeverity", metric.get("baseSeverity", "")),
            "cvss_vector": cvss_data.get("vectorString", ""),
        }
    return {"cvss_version": "", "cvss_score": 0.0, "cvss_severity": "", "cvss_vector": ""}


def parse_nvd_cwes(weaknesses: list[dict[str, Any]]) -> list[str]:
    cwes = []
    for weakness in weaknesses or []:
        for desc in weakness.get("description", []) or []:
            value = desc.get("value", "")
            if value and value.upper().startswith("CWE-"):
                cwes.append(value.upper())
    return dedupe(cwes)


def parse_cveorg_references(items: list[dict[str, Any]], source: str) -> list[dict[str, str]]:
    refs = []
    for item in items or []:
        url = item.get("url", "")
        if not url:
            continue
        refs.append(
            {
                "url": url,
                "source": source,
                "name": item.get("name", ""),
                "tags": ",".join(item.get("tags", []) or []),
            }
        )
    return dedupe_reference_dicts(refs)


def parse_cveorg_affected(items: list[dict[str, Any]]) -> list[str]:
    affected = []
    for item in items or []:
        vendor = item.get("vendor", "")
        product = item.get("product", "")
        label = " ".join(part for part in (vendor, product) if part).strip()
        if label:
            affected.append(label)
    return dedupe(affected)


def parse_cveorg_problem_types(items: list[dict[str, Any]]) -> list[str]:
    values = []
    for item in items or []:
        for desc in item.get("descriptions", []) or []:
            value = desc.get("cweId") or desc.get("description") or ""
            if value:
                values.append(value)
    return dedupe(values)


def first_lang_value(items: list[dict[str, Any]], lang: str) -> str:
    for item in items or []:
        if not lang or item.get("lang") == lang:
            return str(item.get("value") or "").strip()
    return ""


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def extract_urls(text: str) -> list[str]:
    urls = []
    for match in re.finditer(r"https?://[^\s\]\)<>'\"]+", text or ""):
        url = match.group(0).rstrip(".,;:")
        if url:
            urls.append(url)
    return dedupe_urls(urls)


def dedupe_urls(urls: list[str]) -> list[str]:
    result = []
    seen = set()
    for url in urls:
        key = normalize_url_key(url)
        if key and key not in seen:
            seen.add(key)
            result.append(url)
    return result


def dedupe_reference_dicts(refs: list[dict[str, str]]) -> list[dict[str, str]]:
    result = []
    seen = set()
    for ref in refs:
        key = normalize_url_key(ref.get("url", ""))
        if key and key not in seen:
            seen.add(key)
            result.append(ref)
    return result


def normalize_url_key(url: str) -> str:
    return (url or "").strip().rstrip("/").lower()


def dedupe(values: list[Any]) -> list[Any]:
    result = []
    seen = set()
    for value in values:
        key = json.dumps(value, sort_keys=True, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)
        if key not in seen:
            seen.add(key)
            result.append(value)
    return result


def chunks(values: list[str], size: int) -> list[list[str]]:
    return [values[idx : idx + size] for idx in range(0, len(values), size)]


def to_float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def compact_text(text: str, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def compact_product(value: str) -> str:
    text = str(value)
    if text.startswith("cpe:"):
        parts = text.split(":")
        if len(parts) >= 6:
            vendor = parts[3].replace("_", " ")
            product = parts[4].replace("_", " ")
            version = parts[5]
            return " ".join(part for part in (vendor, product, version if version not in {"*", "-"} else "") if part)
    return compact_text(text, 120)


def strip_html(text: str) -> str:
    text = re.sub(r"(?is)<script.*?</script>", " ", text or "")
    text = re.sub(r"(?is)<style.*?</style>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return html_unescape(text)


def strip_markdown(text: str) -> str:
    text = re.sub(r"```.*?```", " ", text or "", flags=re.DOTALL)
    text = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)
    text = re.sub(r"[#>*_`-]+", " ", text)
    return text


def html_unescape(text: str) -> str:
    try:
        import html

        return html.unescape(text)
    except Exception:
        return text


def first_yaml_scalar(text: str, key: str) -> str:
    match = re.search(rf"(?mi)^\s*{re.escape(key)}\s*:\s*(.+?)\s*$", text or "")
    if not match:
        return ""
    return match.group(1).strip().strip('"').strip("'")


def is_rate_limited(response: httpx.Response) -> bool:
    if response.status_code == 429:
        return True
    if response.status_code not in {403, 503}:
        return False
    if response.headers.get("Retry-After"):
        return True
    body = response.text.lower()
    return any(
        marker in body
        for marker in (
            "rate limit",
            "too many requests",
            "quota",
            "exceeded",
            "temporarily unavailable",
        )
    )


def retry_delay(response: httpx.Response, attempt: int) -> int:
    retry_after = response.headers.get("Retry-After")
    if retry_after:
        retry_after = retry_after.strip()
        if retry_after.isdigit():
            return max(1, int(retry_after))
        try:
            retry_at = parsedate_to_datetime(retry_after)
            if retry_at.tzinfo is None:
                retry_at = retry_at.replace(tzinfo=timezone.utc)
            delay = int((retry_at - datetime.now(timezone.utc)).total_seconds())
            return max(1, delay)
        except Exception:
            pass
    return min(30 * attempt, 180)


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


SUMMARY_COLUMNS = [
    ("category", "分类"),
    ("information_score", "信息评分"),
    ("priority_score", "优先级评分"),
    ("cve_id", "CVE编号"),
    ("input_status", "原工作流状态"),
    ("total_reference_count", "总去重参考链接数"),
    ("nvd_reference_count", "NVD引用数"),
    ("cveorg_reference_count", "CVE.org引用数"),
    ("local_reference_count", "本地KB引用数"),
    ("web_result_count", "Web结果数"),
    ("source_hit_count", "来源命中数"),
    ("source_hits", "命中来源"),
    ("poc_hint_count", "PoC线索数"),
    ("cvss_score", "CVSS"),
    ("cvss_severity", "严重性"),
    ("epss", "EPSS"),
    ("epss_percentile", "EPSS百分位"),
    ("kev_known", "CISA KEV"),
    ("published", "发布时间"),
    ("last_modified", "最后修改"),
    ("vuln_status", "状态"),
    ("cwe_ids", "CWE/问题类型"),
    ("affected_count", "受影响产品数"),
    ("affected_sample", "产品样例"),
    ("description_summary", "描述摘要"),
    ("top_references", "主要参考链接"),
    ("suggested_action", "建议动作"),
    ("errors", "错误"),
]


def write_xlsx(records: list[dict[str, Any]], path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "排序汇总"
    write_summary_sheet(summary_sheet, records)
    write_stats_sheet(workbook.create_sheet("分类统计"), records)
    write_reference_sheet(workbook.create_sheet("参考链接明细"), records)
    write_error_sheet(workbook.create_sheet("查询错误"), records)
    workbook.save(path)


def write_summary_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    headers = [label for _, label in SUMMARY_COLUMNS]
    sheet.append(headers)
    for record in records:
        summary = record["summary"]
        row = []
        for key, _label in SUMMARY_COLUMNS:
            if key == "input_status":
                value = record.get("input", {}).get("status", "")
            else:
                value = summary.get(key, "")
            row.append(cell_value(value))
        sheet.append(row)
    style_sheet(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_stats_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    stats: dict[str, int] = {}
    action_stats: dict[str, int] = {}
    for record in records:
        summary = record["summary"]
        stats[summary["category"]] = stats.get(summary["category"], 0) + 1
        action = summary["suggested_action"]
        action_stats[action] = action_stats.get(action, 0) + 1

    sheet.append(["分类", "数量"])
    for category, count in sorted(stats.items()):
        sheet.append([category, count])
    sheet.append([])
    sheet.append(["建议动作", "数量"])
    for action, count in sorted(action_stats.items(), key=lambda item: item[1], reverse=True):
        sheet.append([action, count])
    style_sheet(sheet)


def write_reference_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    sheet.append(["CVE编号", "分类", "信息评分", "来源", "链接", "标签/标题"])
    for record in records:
        summary = record["summary"]
        for ref in collect_references(record):
            sheet.append(
                [
                    record["cve_id"],
                    summary["category"],
                    summary["information_score"],
                    ref.get("source", ""),
                    ref.get("url", ""),
                    compact_text(ref.get("label", ""), 220),
                ]
            )
    style_sheet(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_error_sheet(sheet: Any, records: list[dict[str, Any]]) -> None:
    sheet.append(["CVE编号", "错误"])
    for record in records:
        for error in record["summary"].get("errors", []):
            sheet.append([record["cve_id"], error])
    style_sheet(sheet)


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells[:200]:
            max_len = max(max_len, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 70)


def write_csv(records: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([label for _, label in SUMMARY_COLUMNS])
        for record in records:
            summary = record["summary"]
            row = []
            for key, _label in SUMMARY_COLUMNS:
                if key == "input_status":
                    value = record.get("input", {}).get("status", "")
                else:
                    value = summary.get(key, "")
                row.append(cell_value(value))
            writer.writerow(row)


def write_jsonl(records: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as file:
        for record in records:
            file.write(json.dumps(record, ensure_ascii=False) + "\n")


def cell_value(value: Any) -> str | int | float | bool:
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value or "")


if __name__ == "__main__":
    raise SystemExit(main())
